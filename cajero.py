import tkinter as tk
from tkinter import messagebox, ttk
import mysql.connector
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch

# ------------------ CONEXIÓN ------------------
conexion = mysql.connector.connect(
    host="localhost",
    user="root",
    password="",
    database="billetera"
)

cursor = conexion.cursor()

usuario_actual = None
nombre_actual = None

# ------------------ ESTILO NU -------------------
BG_COLOR = "#4C1D95"
CARD_COLOR = "#5B21B6"
BTN_COLOR = "#9333EA"
BTN2_COLOR = "#7E22CE"
BTN3_COLOR = "#6D28D9"
TEXT_COLOR = "#FFFFFF"

FONT_TITLE = ("Segoe UI", 24, "bold")
FONT_NORMAL = ("Segoe UI", 14)

# ------------------ BALANCE ------------------
def obtener_balance_usuario():
    cursor.execute("""
        SELECT 
        SUM(CASE WHEN tipo='Ingreso' THEN monto ELSE 0 END),
        SUM(CASE WHEN tipo='Gasto' THEN monto ELSE 0 END)
        FROM movimientos
        WHERE usuario_id=%s
    """, (usuario_actual,))
    
    ingresos, gastos = cursor.fetchone()
    ingresos = ingresos or 0
    gastos = gastos or 0
    return ingresos - gastos

# ------------------ SISTEMA PRINCIPAL ------------------
def abrir_sistema():
    ventana = tk.Tk()
    ventana.title("Cuenta Digital")
    ventana.geometry("1000x720")
    ventana.configure(bg=BG_COLOR)

    def cerrar_sesion():
        global usuario_actual, nombre_actual
        usuario_actual = None
        nombre_actual = None
        ventana.destroy()
        abrir_login()

    # Header
    header = tk.Frame(ventana, bg=BG_COLOR)
    header.pack(fill="x", pady=20)

    tk.Label(header,
             text=f"Hola, {nombre_actual}",
             font=FONT_TITLE,
             fg=TEXT_COLOR,
             bg=BG_COLOR).pack(side="left", padx=40)

    tk.Button(header,
              text="Cerrar Sesión",
              bg=BTN_COLOR,
              fg="white",
              font=("Segoe UI", 10, "bold"),
              relief="flat",
              command=cerrar_sesion).pack(side="right", padx=40)

    # Tarjeta balance
    card = tk.Frame(ventana, bg=CARD_COLOR)
    card.pack(pady=20, ipadx=40, ipady=30)

    lbl_balance = tk.Label(card,
                           font=("Segoe UI", 20, "bold"),
                           fg="white",
                           bg=CARD_COLOR)
    lbl_balance.pack()

    # Inputs
    frame_inputs = tk.Frame(ventana, bg=BG_COLOR)
    frame_inputs.pack(pady=20)

    tk.Label(frame_inputs, text="Descripción",
             font=FONT_NORMAL, fg=TEXT_COLOR, bg=BG_COLOR).grid(row=0, column=0, padx=20)

    tk.Label(frame_inputs, text="Monto",
             font=FONT_NORMAL, fg=TEXT_COLOR, bg=BG_COLOR).grid(row=1, column=0, padx=20)

    entry_desc = tk.Entry(frame_inputs, font=FONT_NORMAL, width=25)
    entry_desc.grid(row=0, column=1, pady=10)

    entry_monto = tk.Entry(frame_inputs, font=FONT_NORMAL, width=25)
    entry_monto.grid(row=1, column=1, pady=10)

    def actualizar_balance():
        balance = obtener_balance_usuario()
        lbl_balance.config(text=f"Saldo disponible\n$ {balance:,.2f}")

    def limpiar_campos():
        entry_desc.delete(0, tk.END)
        entry_monto.delete(0, tk.END)

    def agregar_ingreso():
        try:
            monto = float(entry_monto.get())
        except:
            messagebox.showerror("Error", "Monto inválido")
            return

        cursor.execute("""
            INSERT INTO movimientos (usuario_id, tipo, descripcion, monto)
            VALUES (%s,'Ingreso',%s,%s)
        """, (usuario_actual, entry_desc.get(), monto))
        conexion.commit()
        mostrar()
        actualizar_balance()
        limpiar_campos()

    def retirar_dinero():
        try:
            monto = float(entry_monto.get())
        except:
            messagebox.showerror("Error", "Monto inválido")
            return

        balance = obtener_balance_usuario()
        if monto > balance:
            messagebox.showerror("Error", "Saldo insuficiente")
            return

        cursor.execute("""
            INSERT INTO movimientos (usuario_id, tipo, descripcion, monto)
            VALUES (%s,'Gasto',%s,%s)
        """, (usuario_actual, entry_desc.get(), monto))
        conexion.commit()
        mostrar()
        actualizar_balance()
        limpiar_campos()

    def borrar():
        seleccionado = tabla.selection()
        if seleccionado:
            id_mov = tabla.item(seleccionado)['values'][0]
            cursor.execute("DELETE FROM movimientos WHERE id=%s", (id_mov,))
            conexion.commit()
            mostrar()
            actualizar_balance()

    def exportar_excel():
        cursor.execute("""
            SELECT tipo, descripcion, monto, fecha
            FROM movimientos
            WHERE usuario_id=%s
        """, (usuario_actual,))
        datos = cursor.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Movimientos"

        headers = ["Tipo", "Descripción", "Monto", "Fecha"]
        ws.append(headers)

        for col in ws[1]:
            col.font = Font(bold=True, color="FFFFFF")
            col.fill = PatternFill(start_color="4C1D95", end_color="4C1D95", fill_type="solid")
            col.alignment = Alignment(horizontal="center")

        for fila in datos:
            ws.append(fila)

        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 5

        wb.save("movimientos.xlsx")
        messagebox.showinfo("Éxito", "Exportado a Excel correctamente")

    def exportar_pdf():
        cursor.execute("""
            SELECT tipo, descripcion, monto, fecha
            FROM movimientos
            WHERE usuario_id=%s
        """, (usuario_actual,))
        datos = cursor.fetchall()

        doc = SimpleDocTemplate("movimientos.pdf")
        elementos = []

        estilos = getSampleStyleSheet()
        titulo = Paragraph("<b>Estado de Cuenta</b>", estilos["Title"])
        elementos.append(titulo)
        elementos.append(Spacer(1, 0.3 * inch))

        data = [["Tipo", "Descripción", "Monto", "Fecha"]]
        for fila in datos:
            data.append([str(x) for x in fila])

        tabla_pdf = Table(data)
        tabla_pdf.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#4C1D95")),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN',(0,0),(-1,-1),'CENTER'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('FONTSIZE', (0,0), (-1,-1), 9),
        ]))

        elementos.append(tabla_pdf)
        doc.build(elementos)

        messagebox.showinfo("Éxito", "Exportado a PDF correctamente")

    # Botones
    frame_botones = tk.Frame(ventana, bg=BG_COLOR)
    frame_botones.pack(pady=10)

    tk.Button(frame_botones, text="Ingresar Dinero", bg=BTN_COLOR,
              fg="white", font=("Segoe UI", 12, "bold"),
              relief="flat", width=18,
              command=agregar_ingreso).grid(row=0, column=0, padx=15)

    tk.Button(frame_botones, text="Retirar Dinero", bg=BTN2_COLOR,
              fg="white", font=("Segoe UI", 12, "bold"),
              relief="flat", width=18,
              command=retirar_dinero).grid(row=0, column=1, padx=15)

    tk.Button(frame_botones, text="Borrar Movimiento", bg=BTN3_COLOR,
              fg="white", font=("Segoe UI", 12, "bold"),
              relief="flat", width=18,
              command=borrar).grid(row=0, column=2, padx=15)

    tk.Button(frame_botones, text="Exportar a Excel", bg=BTN_COLOR,
              fg="white", font=("Segoe UI", 12, "bold"),
              relief="flat", width=18,
              command=exportar_excel).grid(row=1, column=0, pady=15)

    tk.Button(frame_botones, text="Exportar a PDF", bg=BTN2_COLOR,
              fg="white", font=("Segoe UI", 12, "bold"),
              relief="flat", width=18,
              command=exportar_pdf).grid(row=1, column=1, pady=15)

    # Tabla
    style = ttk.Style()
    style.theme_use("clam")
    style.configure("Treeview",
                    background="white",
                    foreground="black",
                    rowheight=28,
                    fieldbackground="white")

    tabla = ttk.Treeview(ventana,
                         columns=("ID","Usuario","Tipo","Desc","Monto","Fecha"),
                         show="headings",
                         height=12)

    for col in ("ID","Usuario","Tipo","Desc","Monto","Fecha"):
        tabla.heading(col, text=col)
        tabla.column(col, width=150)

    tabla.pack(pady=30)

    def mostrar():
        for row in tabla.get_children():
            tabla.delete(row)

        cursor.execute("""
            SELECT m.id, u.nombre, m.tipo, m.descripcion, m.monto, m.fecha
            FROM movimientos m
            JOIN usuarios u ON m.usuario_id = u.id
            WHERE m.usuario_id=%s
        """, (usuario_actual,))

        for fila in cursor.fetchall():
            tabla.insert("", "end", values=fila)

    actualizar_balance()
    mostrar()
    ventana.mainloop()

# ------------------ LOGIN ---------------------
def abrir_login():
    global ventana_login, entry_nombre, entry_email, entry_password

    ventana_login = tk.Tk()
    ventana_login.title("Acceso Cuenta Digital")
    ventana_login.geometry("550x600")
    ventana_login.configure(bg=BG_COLOR)

    tk.Label(ventana_login, text="Cuenta Digital",
             font=FONT_TITLE, fg="white",
             bg=BG_COLOR).pack(pady=50)

    tk.Label(ventana_login, text="Nombre (Registro)",
             font=FONT_NORMAL, fg="white",
             bg=BG_COLOR).pack()

    entry_nombre = tk.Entry(ventana_login, font=FONT_NORMAL, width=25)
    entry_nombre.pack(pady=10)

    tk.Label(ventana_login, text="Email",
             font=FONT_NORMAL, fg="white",
             bg=BG_COLOR).pack()

    entry_email = tk.Entry(ventana_login, font=FONT_NORMAL, width=25)
    entry_email.pack(pady=10)

    tk.Label(ventana_login, text="Password",
             font=FONT_NORMAL, fg="white",
             bg=BG_COLOR).pack()

    entry_password = tk.Entry(ventana_login, show="*", font=FONT_NORMAL, width=25)
    entry_password.pack(pady=10)

    tk.Button(ventana_login, text="Registrarse",
              bg=BTN_COLOR, fg="white",
              font=("Segoe UI", 12, "bold"),
              relief="flat", command=registrar).pack(pady=20)

    tk.Button(ventana_login, text="Iniciar Sesión",
              bg=BTN2_COLOR, fg="white",
              font=("Segoe UI", 12, "bold"),
              relief="flat", command=login).pack(pady=10)

    ventana_login.mainloop()

def registrar():
    nombre = entry_nombre.get().strip()
    email = entry_email.get().strip()
    password = entry_password.get().strip()

    if not nombre or not email or not password:
        messagebox.showerror("Error", "Todos los campos son obligatorios")
        return

    if len(nombre) < 3 or not nombre.isalpha():
        messagebox.showerror("Error", "Nombre inválido")
        return

    patron_email = r"^[\w\.-]+@[\w\.-]+\.\w+$"
    if not re.match(patron_email, email):
        messagebox.showerror("Error", "Email inválido")
        return

    if len(password) < 6:
        messagebox.showerror("Error", "Contraseña muy corta")
        return

    if not re.search(r"[A-Z]", password) or not re.search(r"[0-9]", password):
        messagebox.showerror("Error", "Debe tener 1 mayúscula y 1 número")
        return

    try:
        cursor.execute("INSERT INTO usuarios (nombre, email, password) VALUES (%s,%s,%s)",
                       (nombre, email, password))
        conexion.commit()
        messagebox.showinfo("Éxito", "Usuario registrado")
    except:
        messagebox.showerror("Error", "Email ya existe")

def login():
    global usuario_actual, nombre_actual

    email = entry_email.get().strip()
    password = entry_password.get().strip()

    cursor.execute("SELECT id, nombre FROM usuarios WHERE email=%s AND password=%s",
                   (email, password))
    resultado = cursor.fetchone()

    if resultado:
        usuario_actual = resultado[0]
        nombre_actual = resultado[1]
        ventana_login.destroy()
        abrir_sistema()
    else:
        messagebox.showerror("Error", "Credenciales incorrectas")

abrir_login()